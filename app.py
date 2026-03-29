import os
import io
import base64
import datetime
import json
from copy import copy
from flask import Flask, request, render_template, send_file, jsonify
import anthropic
import openpyxl
from openpyxl import load_workbook

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max upload

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

EXTRACT_PROMPT = """Extract rental certification data from this PDF and return ONLY a JSON object (no markdown, no explanation) with these exact keys:
{
  "unit_number": "",
  "bedrooms": "",
  "square_feet": "",
  "tc_income_level": "",
  "tc_rent_level": "",
  "tenant_rent": "",
  "initial_cert_date": "MM/DD/YYYY",
  "household_name": "",
  "household_size": "",
  "annual_income": "",
  "pct_ami": "",
  "lease_start_date": "MM/DD/YYYY",
  "security_deposit": ""
}
Return empty string for any field not found. For money fields include $ and commas (e.g. $1,234.56). For percentages include % (e.g. 60%). Lease end date is calculated automatically as 6 months after lease start. The property address is always 1200 Cherry St. Default bedrooms to 2, TC Income Level to 60%, TC Rent Level to 60% if not found."""


def add_months(dt, months):
    month = dt.month - 1 + months
    year = dt.year + month // 12
    month = month % 12 + 1
    day = min(dt.day, [31,28,29,30,31,30,31,31,30,31,30,31][month-1])
    return datetime.datetime(year, month, day)


def extract_from_pdf(pdf_bytes):
    b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")
    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=1000,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "document",
                    "source": {"type": "base64", "media_type": "application/pdf", "data": b64}
                },
                {"type": "text", "text": EXTRACT_PROMPT}
            ]
        }]
    )
    text = "".join(block.text for block in response.content if hasattr(block, "text"))
    clean = text.replace("```json", "").replace("```", "").strip()
    return json.loads(clean)


def update_excel(xlsx_bytes, extracted_rows):
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    # Capture formatting from first data row (row 3)
    def get_fmt(cell):
        return {
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'alignment': copy(cell.alignment),
            'border': copy(cell.border),
            'number_format': cell.number_format,
        }

    ref_formats = {col: get_fmt(ws.cell(row=3, column=col)) for col in range(1, 18)}

    # Collect existing data rows
    existing = [list(row) for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True)]

    # Build new rows from extracted data
    for r in extracted_rows:
        lease_start = None
        lease_end = None
        if r.get("lease_start_date"):
            try:
                lease_start = datetime.datetime.strptime(r["lease_start_date"], "%m/%d/%Y")
                lease_end = add_months(lease_start, 6)
            except Exception:
                pass

        new_row = [
            "1200 Cherry St",
            r.get("unit_number", ""),
            r.get("bedrooms", "2") or "2",
            r.get("square_feet") or None,
            r.get("tc_income_level", "60%") or "60%",
            r.get("tc_rent_level", "60%") or "60%",
            r.get("tenant_rent", ""),
            None,
            r.get("initial_cert_date", ""),
            r.get("household_name", ""),
            r.get("household_size", ""),
            r.get("annual_income", ""),
            r.get("pct_ami", ""),
            None,
            lease_start,
            lease_end,
            r.get("security_deposit", ""),
        ]
        existing.append(new_row)

    # Sort by unit number
    existing.sort(key=lambda row: int(row[1]) if row[1] and str(row[1]).isdigit() else 9999)

    # Delete old data rows and rewrite
    for row_idx in range(ws.max_row, 2, -1):
        ws.delete_rows(row_idx)

    for i, row_data in enumerate(existing):
        excel_row = i + 3
        for col_idx, value in enumerate(row_data):
            col = col_idx + 1
            cell = ws.cell(row=excel_row, column=col)
            cell.value = value
            fmt = ref_formats.get(col, ref_formats[1])
            cell.font = copy(fmt['font'])
            cell.fill = copy(fmt['fill'])
            cell.alignment = copy(fmt['alignment'])
            cell.border = copy(fmt['border'])
            if col in (15, 16) and isinstance(value, datetime.datetime):
                cell.number_format = 'MM/DD/YYYY'
            else:
                cell.number_format = fmt['number_format']

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/process", methods=["POST"])
def process():
    try:
        xlsx_file = request.files.get("xlsx")
        pdf_files = request.files.getlist("pdfs")

        if not xlsx_file:
            return jsonify({"error": "No Excel file uploaded"}), 400
        if not pdf_files:
            return jsonify({"error": "No PDF files uploaded"}), 400

        xlsx_bytes = xlsx_file.read()
        extracted_rows = []
        results = []

        for pdf in pdf_files:
            pdf_bytes = pdf.read()
            try:
                data = extract_from_pdf(pdf_bytes)
                extracted_rows.append(data)
                results.append({
                    "filename": pdf.filename,
                    "status": "success",
                    "unit": data.get("unit_number", "?"),
                    "household": data.get("household_name", "?"),
                })
            except Exception as e:
                results.append({
                    "filename": pdf.filename,
                    "status": "error",
                    "error": str(type(e).__name__) + ": " + str(e)
                })

        if not extracted_rows:
            return jsonify({"error": "No PDFs could be extracted", "results": results}), 400

        updated_xlsx = update_excel(xlsx_bytes, extracted_rows)

        # Store in memory and return as download
        response = send_file(
            updated_xlsx,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="willow_glen_updated.xlsx"
        )
        response.headers["X-Results"] = json.dumps(results)
        return response

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
